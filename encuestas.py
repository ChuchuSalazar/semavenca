import streamlit as st
import pandas as pd
import random
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import streamlit.components.v1 as components

# Función para generar un ID aleatorio


def generar_id():
    return random.randint(100000, 999999)


# Cargar las preguntas del archivo de Excel
df_preguntas = pd.read_excel('preguntas.xlsx')

# Función para guardar las respuestas en el archivo Excel


def guardar_respuestas(respuestas):
    id_encuesta = generar_id()
    fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Crear un diccionario con todas las respuestas y los datos adicionales
    data = {
        'ID': id_encuesta,
        'FECHA': fecha,
        'SEXO': respuestas.get('sexo', ''),
        'RANGO_EDA': respuestas.get('rango_edad', ''),
        'RANGO_INGRESO': respuestas.get('rango_ingreso', ''),
        'CIUDAD': respuestas.get('ciudad', ''),
        'NIVEL_PROF': respuestas.get('nivel_educativo', '')
    }

    # Añadir las respuestas de las preguntas, usando los números correspondientes
    for i, pregunta in enumerate(df_preguntas['item']):
        # Guardar solo el número de la respuesta
        data[f'AV{i+1}'] = respuestas.get(f'AV{i+1}', '')

    # Intentamos guardar las respuestas en el archivo Excel
    try:
        # Intentar cargar el archivo existente
        wb = load_workbook('respuestas.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, creamos uno nuevo
        wb = load_workbook('respuestas.xlsx')
        sheet = wb.active
        # Escribir encabezado
        encabezado = ['ID', 'FECHA', 'SEXO', 'RANGO_EDA', 'RANGO_INGRESO',
                      'CIUDAD', 'NIVEL_PROF'] + [f'AV{i+1}' for i in range(len(df_preguntas))]
        sheet.append(encabezado)

    # Añadir la nueva fila de respuestas
    respuestas_fila = [data['ID'], data['FECHA'], data['SEXO'], data['RANGO_EDA'], data['RANGO_INGRESO'],
                       data['CIUDAD'], data['NIVEL_PROF']] + [data[f'AV{i+1}'] for i in range(len(df_preguntas))]
    sheet.append(respuestas_fila)

    # Guardar los cambios
    wb.save('respuestas.xlsx')

# Función para mostrar la encuesta


def mostrar_encuesta():
    respuestas = {}

    # Mostrar los datos demográficos en forma horizontal
    sexo = st.radio("Sexo:", ['M - Masculino', 'F - Femenino',
                    'O - Otro'], key='sexo', horizontal=True)
    respuestas['sexo'] = sexo.split()[0]  # Guardamos solo el código (M, F, O)

    # Rango de edad
    rango_edad = st.radio("Rango de edad:", [
                          '1 - 18-25', '2 - 26-35', '3 - 36-45', '4 - 46-60', '5 - Más de 60'], key='rango_edad', horizontal=True)
    # Guardamos solo el número del rango
    respuestas['rango_edad'] = rango_edad.split()[0]

    # Rango de ingresos
    rango_ingreso = st.radio("Rango de ingresos (US$):", [
                             '1 - 0-300', '2 - 301-700', '3 - 701-1100', '4 - 1101-1500', '5 - 1501-3000', '6 - Más de 3000'], key='rango_ingreso', horizontal=True)
    # Guardamos solo el número del rango
    respuestas['rango_ingreso'] = rango_ingreso.split()[0]

    # Ciudad como un combo list
    ciudad = st.selectbox("Ciudad:", ['1 - Ciudad A', '2 - Ciudad B',
                          '3 - Ciudad C', '4 - Ciudad D', '5 - Ciudad E'], key='ciudad')
    # Guardamos solo el número de la ciudad
    respuestas['ciudad'] = ciudad.split()[0]

    # Nivel educativo
    nivel_educativo = st.radio("Nivel educativo:", [
                               '1 - Primaria', '2 - Secundaria', '3 - Universitario', '4 - Postgrado'], key='nivel_educativo', horizontal=True)
    # Guardamos solo el número del nivel
    respuestas['nivel_educativo'] = nivel_educativo.split()[0]

    # Mostrar las preguntas numeradas y enmarcadas
    for i, row in df_preguntas.iterrows():
        pregunta_id = row['item']
        pregunta_texto = row['pregunta']
        escala = row['posibles respuestas'].split(',')

        # Numerar y enmarcar las preguntas, además de ajustar la tipografía
        st.markdown(f"**Pregunta {i+1}:**")
        st.markdown(f'<div style="border: 2px solid #add8e6; padding: 10px; border-radius: 5px; font-size: 16px; font-family: Arial, sans-serif;">{
                    pregunta_texto}</div>', unsafe_allow_html=True)

        # Mostrar opciones de respuesta como texto, pero guardar solo el número
        respuesta = st.radio(f"", escala, key=f'AV{pregunta_id}')
        # Guardamos el índice numérico (1, 2, 3, ...)
        respuestas[f'AV{pregunta_id}'] = escala.index(respuesta) + 1

    # Botón para enviar las respuestas
    if st.button("Enviar"):
        # Validar que todas las preguntas hayan sido respondidas
        if all(respuestas.values()):
            guardar_respuestas(respuestas)
            # Mostrar el mensaje de agradecimiento con confeti
            st.balloons()  # Este es el efecto de confeti
            st.success(
                "Gracias por completar la encuesta. ¡Tu respuesta ha sido registrada!")
            st.stop()  # Detener la ejecución para evitar que la encuesta se siga mostrando
        else:
            st.error("Por favor, responde todas las preguntas.")


# Llamar la función para mostrar la encuesta
if __name__ == '__main__':
    mostrar_encuesta()
